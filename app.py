from flask import Flask, render_template, request, jsonify
import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from flask import send_file
import io

app = Flask(__name__)

DATA_FILE = 'transactions.json'

def load_transactions():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []

def save_transactions(transactions):
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(transactions, f, ensure_ascii=False, indent=2)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/transactions', methods=['GET'])
def get_transactions():
    transactions = load_transactions()
    return jsonify(transactions)

@app.route('/api/transactions', methods=['POST'])
def add_transaction():
    data = request.json
    transactions = load_transactions()
    
    new_transaction = {
        'id': max([t['id'] for t in transactions], default=0) + 1 if transactions else 1,
        'type': data['type'],
        'amount': float(data['amount']),
        'category': data['category'],
        'description': data['description'],
        'date': data['date'],
        'timestamp': datetime.now().isoformat()
    }
    
    transactions.insert(0, new_transaction)
    save_transactions(transactions)
    
    return jsonify(new_transaction), 201

@app.route('/api/transactions/<int:id>', methods=['DELETE'])
def delete_transaction(id):
    transactions = load_transactions()
    transactions = [t for t in transactions if t['id'] != id]
    save_transactions(transactions)
    return jsonify({'success': True})

@app.route('/api/transactions/<int:id>', methods=['PUT'])
def update_transaction(id):
    data = request.json
    transactions = load_transactions()
    
    for i, t in enumerate(transactions):
        if t['id'] == id:
            transactions[i] = {
                'id': id,
                'type': data['type'],
                'amount': float(data['amount']),
                'category': data['category'],
                'description': data['description'],
                'date': data['date'],
                'timestamp': t.get('timestamp', datetime.now().isoformat())
            }
            break
    
    save_transactions(transactions)
    return jsonify({'success': True})


from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from flask import send_file
import io

@app.route('/api/transactions/export/excel')
def export_excel():
    transactions = load_transactions()

    wb = Workbook()
    ws = wb.active
    ws.title = "Bütçe Kayıtları"

    headers = [
        'ID',
        'Tür',
        'Tutar (₺)',
        'Kategori',
        'Açıklama',
        'Tarih'
    ]

    header_fill = PatternFill("solid", fgColor="6D28D9")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center")

    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col)].width = 20

    for t in transactions:
        ws.append([
            t['id'],
            'Gelir' if t['type'] == 'income' else 'Gider',
            t['amount'],
            t['category'],
            t['description'],
            t['date']
        ])

    # Gelir / gider renklendirme
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            tur = ws.cell(row=cell.row, column=2).value
            cell.font = Font(
                color="16A34A" if tur == 'Gelir' else "DC2626",
                bold=True
            )

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name="butce_raporu.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


if __name__ == '__main__':
    app.run(debug=True, port=5000)